import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook
import time
import pandas as pd

time_now = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())

class AADBFolder(data.Dataset):

    def __init__(self, istrain, root, transform, patch_num):
        imgname = []
        mos_all = []
        if istrain:
            csv_file = os.path.join(root, 'imgListTrainRegression_score.txt')
        else:
            csv_file = os.path.join(root, 'imgListTestNewRegression_score.txt')

        with open(csv_file) as f:
            for line in f.readlines():
                temp = line.strip().split()
                imgname.append(temp[0])
                mos = np.array(float(temp[1])*10).astype(np.float32)
                mos_all.append(mos)
        print(imgname[0])
        print(mos_all[0])


        sample = []
        for i, item in enumerate(imgname):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'original', item), mos_all[i]))
        print(len(sample))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class TestFolder(data.Dataset):
    def __init__(self, transform, patch_num):
        test_root = '/usr/zhengjia/deblurred_images'
        test_file = getFileName(test_root, '.png')
        test_file_PNG = getFileName(test_root, '.PNG')
        test_file_JPG = getFileName(test_root, '.JPG')
        test_file = test_file_PNG + test_file + test_file_JPG
        print(test_file)
        print(len(test_file))

        sample = []
        for i in range(len(test_file)):
            for aug in range(patch_num):
                sample.append(os.path.join('/usr/zhengjia/deblurred_images',test_file[i]))

        print('长度', len(sample))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample

    def __len__(self):
        length = len(self.samples)
        return length


class AVAFolder(data.Dataset):

    def __init__(self,istrain, root, index, transform, patch_num):

        images_inf = pd.read_csv('my_train_and_test.csv')
        # Index([ 'MOS', 'image_name', 'MLS','set_new', 'width', 'height', 'ID', 'threshold', 'mos_float'],
        sample = []
        if istrain:
            for i, item in enumerate(list(images_inf['set_new'])):
                if item == 'train':
                    for aug in range(patch_num):
                        lists = (images_inf['1'][i], images_inf['2'][i], images_inf['3'][i], images_inf['4'][i], images_inf['5'][i],
                                images_inf['6'][i], images_inf['7'][i], images_inf['8'][i], images_inf['9'][i], images_inf['10'][i])
                        sample.append((os.path.join(root,images_inf['image_name'][i]), lists,
                                      np.array(images_inf['mos_float'][i]).astype(np.float32)))
                # if i == 10000:
                #     break
            print('train数据：', len(sample))
        else:
            for i, item in enumerate(list(images_inf['set_new'])):
                if item == 'test':
                    for aug in range(patch_num):
                        lists = (
                        images_inf['1'][i], images_inf['2'][i], images_inf['3'][i], images_inf['4'][i], images_inf['5'][i],
                        images_inf['6'][i], images_inf['7'][i], images_inf['8'][i], images_inf['9'][i], images_inf['10'][i])
                        sample.append((os.path.join(root, images_inf['image_name'][i]), lists,
                                      np.array(images_inf['mos_float'][i]).astype(np.float32)))
                # if i == 10000:
                #     break
            print('test数据：', len(sample))
        # for i, item in enumerate(index):
        #     for aug in range(patch_num):
        #         sample.append((os.path.join(root, item[0]),item[1], np.array(item[2]).astype(np.float32)))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        #num = list(range(0,len(index)))
        path,p_tar, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, p_tar, target

    def __len__(self):
        length = len(self.samples)
        return length

class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)  # 获得每一张图像的路径
        # [0:227] [227:460] [460:634] [634:808] [808:982] 分别为五种失真类型的索引，可单独训练每一种失真类型
        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname
        imgpath = imgpath
        dmos = scipy.io.loadmat(os.path.join(root, 'dmos.mat'))
        labels = dmos['dmos'].astype(np.float32)
        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']
        sample = []

        print(refname)
        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((imgpath[item], labels[0][item]))
        print('长度', len(sample))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []
        #csv_file = os.path.join('/userHOME/zj/dataset/ChallengeDB_release', 'good.csv')  # good quality
        csv_file = os.path.join('/userHOME/zj/dataset/ChallengeDB_release', 'bad.csv')  # bad quality
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['imgpath'])
                mos = np.array(float(row['mos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'Images', imgname[item]), mos_all[item]))
        print("liveC:", len(sample))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class CSIQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath, '.png')
        txtpath = os.path.join(root, 'csiq_label_jpeg2000.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'dst_imgs_all', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform
        print(len(self.samples))
        # np.savetxt('./{}train_pair.txt'.format(time_now), self.samples , fmt="%s", delimiter=" ")

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq_10kFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        imgname = []
        mos_all = []
        #csv_file = os.path.join('/userHOME/zj/dataset/koniq10k', 'good.csv')  # good quality
        csv_file = os.path.join('/userHOME/zj/dataset/koniq10k', 'bad.csv')  # bad quality
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['imgpath'])
                mos = np.array(float(row['mos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, '1024x768', imgname[item]), mos_all[item]))
        print(len(sample))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class BIDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'Image', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


# class TID2013Folder(data.Dataset):
#
#     def __init__(self, root, index, transform, patch_num):
#         refpath = os.path.join(root, 'reference_images')
#         refname = getTIDFileName(refpath, '.bmp.BMP')
#         txtpath = os.path.join(root, 'mos_with_names.txt')
#         fh = open(txtpath, 'r')
#         imgnames = []
#         target = []
#         refnames_all = []
#         for line in fh:
#             line = line.split('\n')
#             words = line[0].split()
#             imgnames.append((words[1]))
#             target.append(words[0])
#             ref_temp = words[1].split("_")
#             refnames_all.append(ref_temp[0][1:])
#         labels = np.array(target).astype(np.float32)
#         refnames_all = np.array(refnames_all)
#
#         sample = []
#         for i, item in enumerate(index):
#             train_sel = (refname[index[i]] == refnames_all)
#             train_sel = np.where(train_sel == True)
#             train_sel = train_sel[0].tolist()
#             for j, item in enumerate(train_sel):
#                 for aug in range(patch_num):
#                     sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
#         print("tid2013", len(sample))
#         self.samples = sample
#         self.transform = transform
#
#     def __getitem__(self, index):
#         """
#         Args:
#             index (int): Index
#
#         Returns:
#             tuple: (sample, target) where target is class_index of the target class.
#         """
#         path, target = self.samples[index]
#         sample = pil_loader(path)
#         sample = self.transform(sample)
#         return sample, target
#
#     def __len__(self):
#         length = len(self.samples)
#         return length
class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath, '.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []

        target = []
        refnames_all = []
        refnames_all2 = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
            refnames_all2.append(ref_temp[0] + "_" + ref_temp[1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)
        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)  # 返回参考图像对应的失真图的全部索引
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(1):
                    sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
        print("zongchangdu:", len(sample))
        # print(sample)
        # 0:5 表示每个参考图对应的第一种失真类型 单独训练每一种失真类型  下面的分段表示没一种失真类型，一种24种
        # 0:5  5:10  10:15  15:20 20:25 25:30 30:35 35:40 40:45 45:50 50:55
        # 55:60 60:65  65:70  70:75  75:80  80:85  85:90  90:95  95:100  100:105
        # 105:110  110:115  115:120
        sample = np.array(sample)
        imgnames2 = sample[:, 0].reshape(len(index), 120)[:, 105:110]
        labels2 = sample[:, 1].reshape(len(index), 120)[:, 105:110]
        imgnames2 = imgnames2.reshape(imgnames2.size, ).tolist()
        labels2 = labels2.reshape(labels2.size, ).tolist()

        print('qiepianzhihou:', len(labels2))
        temp = []
        for i in range(len(labels2)):
            for _ in range(patch_num):
                temp.append((imgnames2[i], np.float32(labels2[i])))
        print(temp)
        print('zengqiangzhihou:', len(temp))
        self.samples = temp
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')
